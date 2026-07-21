"""Аналіз цін між нашим магазином (Агродрузі) та конкурентами.

Логіка (див. README.md):

1. Для кожного рядка Input.xlsx беремо значення колонки "Артикул" і шукаємо
   співпадіння в таблицях конкурентів (agrodoctor_prices.xlsx, tvk_ramos_prices.xlsx).
   1.1 Шукаємо співпадіння в колонках name, references, description.
   1.2 Якщо повного співпадіння немає і артикул містить крапку "." — шукаємо по
       значенню артикула до крапки (напр. "AZ10331.41" -> "AZ10331").
2. Для кожного джерела окремо додаємо колонку з найдешевшою знайденою ціною:
   "Ціна Рамос" (tvk-ramos) та "Ціна Агродоктор" (agrodoctor). Якщо в джерелі
   товар не знайдено — ціна 0.
3. Колір конкретної клітинки з ціною:
   - зелений (#82d200), якщо ціна нижча за нашу і товар є в наявності;
   - червоний (#ff6f6d), якщо немає в наявності, ціна вища/рівна або не знайдено.
4. Якщо ціна знайдена (!= 0) — вона стає клікабельним посиланням на товар.

Назва результуючого файлу містить дату аналізу у форматі YYYY-MM-DD.
"""

import re
from collections import defaultdict
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# --- Конфігурація ---------------------------------------------------------

INPUT_FILE = 'Input.xlsx'
# До назви результуючого файлу дописуємо дату у форматі YYYY-MM-DD.
OUTPUT_FILE = f'Input_analyzed_with_urls_{date.today().isoformat()}.xlsx'

COMPETITORS = [
    ('agrodoctor', 'agrodoctor_prices.xlsx'),
    ('tvk-ramos', 'tvk_ramos_prices.xlsx'),
]

# Колонки конкурентів, у яких шукаємо артикул.
SEARCH_COLUMNS = ['name', 'references', 'description']

# Колонки нашого файлу.
COL_ARTICLE = 'Артикул'
COL_OUR_PRICE = 'Ціна Агродрузі'

# Для кожного конкурента — своя колонка з його ціною (порядок = порядок колонок).
SOURCE_COLUMNS = [
    ('tvk-ramos', 'Ціна Рамос'),
    ('agrodoctor', 'Ціна Агродоктор'),
]

# Ціна-заглушка, коли товар у джерелі не знайдено.
PRICE_NOT_FOUND = 0

OUT_OF_STOCK = 'Немає в наявності'

# Кольори (ARGB для openpyxl).
FILL_GREEN = PatternFill('solid', fgColor='FF82D200')  # ціна нижча за нашу і є в наявності
FILL_RED = PatternFill('solid', fgColor='FFFF6F6D')    # дорожче / немає в наявності / не знайдено

# Шрифт для клікабельних цін-гіперпосилань (підкреслений, читабельний на заливці).
LINK_FONT = Font(underline='single')

# Мінімальна довжина "базового" (до крапки) артикула. Захищає від надто
# загальних токенів на кшталт "38" (з "38.4V/2K1/JA/J2A"), які дали б тисячі
# хибних співпадінь.
MIN_BASE_LEN = 3


# --- Допоміжні функції ----------------------------------------------------

def clean_price(price):
    """Конвертація ціни у число. Непарсабельне значення -> inf."""
    if pd.isna(price):
        return float('inf')
    if isinstance(price, (int, float)):
        return float(price)
    price_str = str(price).replace(' ', '').replace('\xa0', '').replace(',', '.')
    try:
        return float(price_str)
    except ValueError:
        return float('inf')


def split_articles(raw):
    """Один рядок 'Артикул' може містити кілька артикулів через кому/крапку з комою."""
    if pd.isna(raw):
        return []
    text = str(raw).replace(';', ',')
    return [part.strip(' *') for part in text.split(',') if part.strip(' *')]


def base_article(article):
    """Значення артикула до останньої крапки (п.1.2 README).

    Повертає None, якщо крапки немає або база надто коротка/загальна.
    """
    if '.' not in article:
        return None
    base = article.rsplit('.', 1)[0].strip()
    if len(base) < MIN_BASE_LEN or base == article:
        return None
    return base


def combined_text(df):
    """Обʼєднаний текст пошукових колонок для кожного рядка конкурента."""
    parts = [df[col].fillna('').astype(str) for col in SEARCH_COLUMNS]
    joined = parts[0]
    for extra in parts[1:]:
        joined = joined + ' ' + extra
    return joined.tolist()


# --- Основна логіка -------------------------------------------------------

def build_token_map(df_input):
    """token -> список (row_idx, article_idx, level).

    level 0 = повне співпадіння артикула, level 1 = співпадіння по базі (до крапки).
    Рішення "повне чи база" приймається окремо для кожного артикула.
    """
    token_map = defaultdict(list)
    row_art_count = {}

    for row_idx, raw in enumerate(df_input[COL_ARTICLE]):
        articles = split_articles(raw)
        row_art_count[row_idx] = len(articles)
        for art_idx, article in enumerate(articles):
            token_map[article].append((row_idx, art_idx, 0))
            base = base_article(article)
            if base:
                token_map[base].append((row_idx, art_idx, 1))

    return token_map, row_art_count


def build_matcher(tokens):
    """Один регекс на всі токени. Довші — раніше, щоб префікс не "з'їдав" повний.

    Межі \\w з обох боків: артикул має бути цілим токеном, а не частиною довшого
    числа/слова (напр. "560210" не має ловитись усередині "0005602100").
    """
    ordered = sorted(tokens, key=len, reverse=True)
    pattern = r'(?<!\w)(?:' + '|'.join(re.escape(t) for t in ordered) + r')(?!\w)'
    return re.compile(pattern)


def collect_hits(token_map, matcher):
    """Прохід по всіх конкурентах.

    Повертає (source, row_idx, art_idx, level) -> [(price, availability, url), ...].
    Ключ містить source, бо ціни рахуємо окремо для кожного джерела.
    """
    hits = defaultdict(list)

    for source, filename in COMPETITORS:
        df = pd.read_excel(filename)
        texts = combined_text(df)
        prices = df['price'].tolist()
        avails = df['availability'].tolist()
        urls = df['product_url'].tolist()

        for text, price, avail, url in zip(texts, prices, avails, urls):
            price = clean_price(price)
            if price == float('inf'):
                continue
            matched = {m.group(0) for m in matcher.finditer(text)}
            if not matched:
                continue
            candidate = (price, avail, url)
            for token in matched:
                for row_idx, art_idx, level in token_map.get(token, ()):
                    hits[(source, row_idx, art_idx, level)].append(candidate)

    return hits


def best_for_source(hits, source, row_idx, art_count):
    """Найнижча ціна конкретного джерела для рядка. -> (price, availability, url) або None.

    Для кожного артикула рядка: повні співпадіння, або (якщо їх нема) — база (п.1.2).
    """
    pool = []
    for art_idx in range(art_count):
        full = hits.get((source, row_idx, art_idx, 0), [])
        base = hits.get((source, row_idx, art_idx, 1), [])
        pool.extend(full if full else base)
    if not pool:
        return None
    return min(pool, key=lambda c: c[0])


def analyze(df_input, hits, row_art_count):
    """Для кожного джерела формує колонку з ціною + дані для заливки/посилань.

    Повертає col_name -> {'urls': [...], 'fills': [...]}.
    """
    results = {}

    for source, col_name in SOURCE_COLUMNS:
        values, urls, fills = [], [], []

        for row_idx in range(len(df_input)):
            our_price = clean_price(df_input.at[row_idx, COL_OUR_PRICE])
            best = best_for_source(hits, source, row_idx, row_art_count[row_idx])

            if best is None:
                # Товар не знайдено -> ціна 0, клітинка червона.
                values.append(PRICE_NOT_FOUND)
                urls.append(None)
                fills.append(FILL_RED)
                continue

            price, avail, url = best
            in_stock = str(avail).strip() != OUT_OF_STOCK
            # Зелена лише якщо є в наявності і дешевше за нашу; інакше червона.
            green = in_stock and price < our_price
            values.append(price)
            urls.append(None if pd.isna(url) else str(url))
            fills.append(FILL_GREEN if green else FILL_RED)

        df_input[col_name] = values
        results[col_name] = {'urls': urls, 'fills': fills}

    return results


def write_output(df_input, results):
    """Записує результат: заливка клітинок з цінами + гіперпосилання на товар."""
    df_input.to_excel(OUTPUT_FILE, index=False)

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    for col_name, data in results.items():
        col_idx = df_input.columns.get_loc(col_name) + 1  # openpyxl рахує з 1
        for row_idx, (url, fill) in enumerate(zip(data['urls'], data['fills'])):
            cell = ws.cell(row=row_idx + 2, column=col_idx)  # +1 заголовок, +1 зсув
            cell.fill = fill
            # Ціна != 0 і є посилання -> робимо її клікабельною.
            if url and cell.value != PRICE_NOT_FOUND:
                cell.hyperlink = url
                cell.font = LINK_FONT
    wb.save(OUTPUT_FILE)


def main():
    df_input = pd.read_excel(INPUT_FILE)

    token_map, row_art_count = build_token_map(df_input)
    matcher = build_matcher(token_map.keys())
    hits = collect_hits(token_map, matcher)
    results = analyze(df_input, hits, row_art_count)
    write_output(df_input, results)

    print(f'Аналіз завершено. Результати збережено у {OUTPUT_FILE}')
    for _, col_name in SOURCE_COLUMNS:
        fills = results[col_name]['fills']
        green = sum(1 for f in fills if f is FILL_GREEN)
        print(f'  {col_name}: дешевше за нас (зелені) — {green} з {len(fills)}')


if __name__ == '__main__':
    main()
